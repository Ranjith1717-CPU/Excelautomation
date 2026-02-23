"""
=============================================================================
FORMATTER MODULE
=============================================================================
Excel visual formatting: charts, colors, table styles, column sizing.

Functions:
  add_bar_chart          - Embedded bar chart
  add_line_chart         - Embedded line chart
  add_pie_chart          - Embedded pie chart
  apply_traffic_light    - Red/Yellow/Green cell fills
  apply_color_scale      - Gradient min→max color scale
  format_as_table        - Apply Excel table style
  freeze_and_filter      - Freeze header row + auto-filter
  auto_fit_columns       - Auto-width based on content
  add_totals_row         - SUM row at bottom for numeric columns
  highlight_duplicates   - Color duplicate cells yellow
  apply_number_format    - Currency/percentage/comma format
=============================================================================
"""
import pandas as pd
import numpy as np
from pathlib import Path
from typing import List, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


# ── helpers ──────────────────────────────────────────────────────────────────

def _load_wb(file: str):
    """Load workbook + active sheet."""
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    return wb, ws


def _save_wb(wb, output_path: str) -> str:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"    Saved   : {output_path}")
    return output_path


def _col_idx(ws, col_name: str) -> int:
    """Find 1-based column index by header name in row 1."""
    for cell in ws[1]:
        if str(cell.value) == col_name:
            return cell.column
    raise ValueError(f"Column '{col_name}' not found in sheet headers")


# ── Charts ────────────────────────────────────────────────────────────────────

def add_bar_chart(file: str, x_col: str, y_col: str,
                  output_path: str, title: str = "Bar Chart") -> str:
    """Add an embedded bar chart to the worksheet."""
    wb, ws = _load_wb(file)
    max_row = ws.max_row
    x_idx = _col_idx(ws, x_col)
    y_idx = _col_idx(ws, y_col)

    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = y_col
    chart.x_axis.title = x_col

    data = Reference(ws, min_col=y_idx, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=x_idx, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4

    anchor_col = get_column_letter(ws.max_column + 2)
    ws.add_chart(chart, f"{anchor_col}2")
    return _save_wb(wb, output_path)


def add_line_chart(file: str, x_col: str, y_col: str,
                   output_path: str, title: str = "Line Chart") -> str:
    """Add an embedded line chart to the worksheet."""
    wb, ws = _load_wb(file)
    max_row = ws.max_row
    x_idx = _col_idx(ws, x_col)
    y_idx = _col_idx(ws, y_col)

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_col
    chart.x_axis.title = x_col

    data = Reference(ws, min_col=y_idx, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=x_idx, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    anchor_col = get_column_letter(ws.max_column + 2)
    ws.add_chart(chart, f"{anchor_col}2")
    return _save_wb(wb, output_path)


def add_pie_chart(file: str, category_col: str, value_col: str,
                  output_path: str, title: str = "Pie Chart") -> str:
    """Add an embedded pie chart to the worksheet."""
    wb, ws = _load_wb(file)
    max_row = ws.max_row
    cat_idx = _col_idx(ws, category_col)
    val_idx = _col_idx(ws, value_col)

    chart = PieChart()
    chart.title = title

    data = Reference(ws, min_col=val_idx, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=cat_idx, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    anchor_col = get_column_letter(ws.max_column + 2)
    ws.add_chart(chart, f"{anchor_col}2")
    return _save_wb(wb, output_path)


# ── Conditional Formatting (manual cell-fill via openpyxl) ───────────────────

def apply_traffic_light(file: str, column: str, output_path: str,
                        red: Optional[float] = None,
                        yellow: Optional[float] = None) -> str:
    """
    Apply Red/Yellow/Green fills to a numeric column.
    Values < red → Red, red ≤ val < yellow → Yellow, else → Green.
    If thresholds not provided, uses 33rd/66th percentile.
    """
    df = pd.read_excel(file)
    wb, ws = _load_wb(file)
    col_idx = _col_idx(ws, column)

    vals = pd.to_numeric(df[column], errors="coerce").dropna()
    if red is None:   red = float(vals.quantile(0.33))
    if yellow is None: yellow = float(vals.quantile(0.66))

    RED_FILL    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    GREEN_FILL  = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        try:
            v = float(cell.value)
            if v < red:      cell.fill = RED_FILL
            elif v < yellow: cell.fill = YELLOW_FILL
            else:             cell.fill = GREEN_FILL
        except (TypeError, ValueError):
            pass

    return _save_wb(wb, output_path)


def apply_color_scale(file: str, column: str, output_path: str) -> str:
    """
    Gradient color scale (white→blue) from min to max in a column.
    """
    df = pd.read_excel(file)
    wb, ws = _load_wb(file)
    col_idx = _col_idx(ws, column)

    vals = pd.to_numeric(df[column], errors="coerce")
    min_val = vals.min()
    max_val = vals.max()
    rng = max_val - min_val if max_val != min_val else 1

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        try:
            v = float(cell.value)
            intensity = int((v - min_val) / rng * 200)  # 0-200
            r = max(0, 255 - intensity)
            g = max(0, 255 - intensity)
            b = 255
            hex_color = f"{r:02X}{g:02X}{b:02X}"
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        except (TypeError, ValueError):
            pass

    return _save_wb(wb, output_path)


# ── Table & Layout ────────────────────────────────────────────────────────────

def format_as_table(file: str, output_path: str,
                    style: str = "TableStyleMedium9") -> str:
    """Apply an Excel table style to the used range."""
    wb, ws = _load_wb(file)
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = openpyxl.worksheet.table.Table(displayName="DataTable", ref=ref)
    table.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
        name=style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    return _save_wb(wb, output_path)


def freeze_and_filter(file: str, output_path: str) -> str:
    """Freeze the header row and enable auto-filter."""
    wb, ws = _load_wb(file)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return _save_wb(wb, output_path)


def auto_fit_columns(file: str, output_path: str,
                     min_width: int = 8, max_width: int = 50) -> str:
    """Set column widths based on the longest content in each column."""
    wb, ws = _load_wb(file)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))
    return _save_wb(wb, output_path)


def add_totals_row(file: str, output_path: str) -> str:
    """Add a SUM totals row at the bottom for all numeric columns."""
    wb, ws = _load_wb(file)
    max_row = ws.max_row
    max_col = ws.max_column

    TOTAL_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    TOTAL_FONT = Font(bold=True, color="FFFFFF")

    ws.cell(row=max_row + 1, column=1).value = "TOTAL"
    ws.cell(row=max_row + 1, column=1).font = TOTAL_FONT
    ws.cell(row=max_row + 1, column=1).fill = TOTAL_FILL

    for col in range(1, max_col + 1):
        cell = ws.cell(row=max_row + 1, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        # Check if column is numeric
        try:
            vals = [ws.cell(row=r, column=col).value for r in range(2, max_row + 1)]
            numeric_vals = [float(v) for v in vals if v is not None and str(v).replace(".", "").replace("-", "").isdigit()]
            if numeric_vals and col > 1:
                col_letter = get_column_letter(col)
                cell.value = f"=SUM({col_letter}2:{col_letter}{max_row})"
        except Exception:
            pass

    return _save_wb(wb, output_path)


def highlight_duplicates(file: str, column: str, output_path: str) -> str:
    """Highlight duplicate values in a column with yellow fill."""
    df = pd.read_excel(file)
    wb, ws = _load_wb(file)
    col_idx = _col_idx(ws, column)

    dup_mask = df[column].duplicated(keep=False)
    YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_idx, is_dup in enumerate(dup_mask, start=2):
        if is_dup:
            ws.cell(row=row_idx, column=col_idx).fill = YELLOW

    return _save_wb(wb, output_path)


def apply_number_format(file: str, columns: List[str],
                        fmt: str, output_path: str) -> str:
    """
    Apply a number format string to specified columns.
    fmt examples:
      '#,##0.00'         → comma thousands with 2 decimals
      '"$"#,##0.00'      → currency
      '0.00%'            → percentage
      '#,##0'            → integer with comma
    """
    wb, ws = _load_wb(file)
    for col_name in columns:
        try:
            col_idx = _col_idx(ws, col_name)
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = fmt
        except ValueError as e:
            print(f"    Warning : {e}")
    return _save_wb(wb, output_path)
